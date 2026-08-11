"""
agents/eval/qa_merge.py
QA셋(시험지) ↔ 외부 로그(답안지) 병합 유틸

상대는 QA셋을 로그와 별도 파일로 준다(로그 JSONL 에 정답을 손수 끼워 넣어주지
않는다). 이 유틸이 질문 매칭으로 QA셋의 정답(ground_truth)·근거 문단
(gold_contexts)을 로그 레코드에 병합해, 스키마 v1 의 "2차 심화" 재료를 만든다
(docs/external_rag_log_intake.md §3).

규칙:
- 매칭: 질문 문자열 정규화(공백 압축·끝 문장부호 제거·소문자) 후 일치.
  못 맞춘 항목은 버리지 않고 집계만 한다(폴백 철학).
- 채우기만 한다: 로그에 이미 있는 ground_truth/gold_contexts 는 덮지 않고
  충돌로 집계(로그 제공자가 직접 넣은 값을 신뢰).
- 키 관용: QA셋 형식은 팀마다 달라서 흔한 키 이름을 폭넓게 받는다. 주의 -
  QA셋의 "answer"/"contexts" 는 시험지 의미(정답/근거)라 로그의 같은 이름
  (시스템 출력/검색 결과)과 반대다. 병합 결과에서는 ground_truth/gold_contexts
  로만 쓴다.
- 깨진 로그 줄은 그대로 통과시킨다(적재기 log_intake 가 일관되게 집계하도록).
- 골든셋 형식은 팀마다 다르므로 JSON 배열 / JSONL / CSV / XLSX 를 모두 받는다
  (멘토 피드백: "골든셋(JSONL, Excel 등)을 입력"). 확장자로 1차 판별하되 실패하면
  JSON→JSONL 순으로 재시도한다 - 상대가 .txt 로 준 JSONL 같은 케이스를 살린다.

CLI: python -m agents.eval.qa_merge <log.jsonl> <qa.json|jsonl|csv|xlsx> [--out=경로]
     기본 출력: <log>.merged.jsonl
"""

from __future__ import annotations

import csv
import io
import json
import os
import re
import sys
from typing import Any, Optional

# QA셋에서 받아들이는 키 이름(앞에 있을수록 우선)
QUESTION_KEYS = ("question", "q", "query")
GROUND_TRUTH_KEYS = ("ground_truth", "gold_answer", "answer", "answers")
GOLD_CONTEXT_KEYS = ("gold_contexts", "gold_context", "evidence", "contexts", "context")

_WS = re.compile(r"\s+")
_TRAIL_PUNCT = re.compile(r"[?？.!\s]+$")


def normalize_question(text: str) -> str:
    """질문 매칭 키 - 공백 압축, 끝 문장부호 제거, 소문자."""
    text = _WS.sub(" ", str(text or "").strip())
    return _TRAIL_PUNCT.sub("", text).lower()


def _first_key(obj: dict, keys: tuple[str, ...]) -> Any:
    for k in keys:
        if k in obj and obj[k] not in (None, "", []):
            return obj[k]
    return None


def _as_text(value: Any) -> Optional[str]:
    """정답 후보 → 텍스트 하나. 리스트(KorQuAD식 복수 정답)는 첫 항목(v1 규약).
    항목이 dict(KorQuAD 원본 answers=[{"text":...,"answer_start":...}])면 그 dict 를
    str() 통짜 변환하지 않고 "text" 키를 꺼낸다 - 안 그러면 "{'text': ..., ...}" 가
    정답이 돼 채점을 조용히 오염시킨다."""
    if isinstance(value, list):
        value = value[0] if value else None
    if isinstance(value, dict):
        value = value.get("text") or value.get("answer")
    text = str(value or "").strip()
    return text or None


def _as_text_list(value: Any) -> list[str]:
    if isinstance(value, str):
        value = [value]
    if not isinstance(value, list):
        return []
    return [t for t in (str(v or "").strip() for v in value) if t]


def _entries_from_json(content: str) -> tuple[list[Any], list[str]]:
    """JSON 배열([{...}]) 또는 최상위 dict({"data": [...]} 류)에서 항목 목록을 뽑는다."""
    errors: list[str] = []
    loaded = json.loads(content)          # JSONDecodeError 는 호출자가 처리
    if isinstance(loaded, list):
        return loaded, errors
    if isinstance(loaded, dict):
        entries = next((v for v in loaded.values() if isinstance(v, list)), [])
        if not entries:
            errors.append("JSON 오브젝트 안에 항목 리스트가 없음")
        return entries, errors
    errors.append("JSON 최상위가 배열/오브젝트가 아님")
    return [], errors


def _entries_from_jsonl(content: str) -> tuple[list[Any], list[str]]:
    entries: list[Any] = []
    errors: list[str] = []
    for lineno, line in enumerate(content.splitlines(), start=1):
        line = line.strip()
        if not line:
            continue
        try:
            entries.append(json.loads(line))
        except json.JSONDecodeError as exc:
            errors.append(f"{lineno}행: {exc}")
    return entries, errors


def _entries_from_csv(content: str) -> tuple[list[Any], list[str]]:
    """CSV → dict 리스트. 헤더가 곧 키라서 키 관용 규칙(QUESTION_KEYS 등)이 그대로 통한다.
    빈 셀은 키를 아예 빼서 _first_key 의 '값 없으면 다음 후보' 규칙을 타게 한다."""
    rows = list(csv.DictReader(io.StringIO(content)))
    entries = [{k.strip(): v for k, v in row.items()
                if k and str(v or "").strip()} for row in rows]
    return entries, []


def _entries_from_xlsx(path: str) -> tuple[list[Any], list[str]]:
    """XLSX 첫 시트 → dict 리스트(1행=헤더). openpyxl 미설치면 오류로 알린다."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], ["xlsx 를 읽으려면 openpyxl 이 필요합니다 (pip install openpyxl)"]
    # read_only 는 파일 핸들을 열어둔 채라 Windows 에서 원본을 못 지운다 - 반드시 close.
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        rows = wb.active.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(rows, [])]
        entries = []
        for row in rows:
            entry = {h: v for h, v in zip(header, row)
                     if h and str(v or "").strip()}
            if entry:
                entries.append(entry)
    finally:
        wb.close()
    return entries, []


def _read_text(path: str) -> tuple[Optional[str], list[str]]:
    """utf-8-sig(BOM 붙은 엑셀 CSV 수용) 로 먼저 읽고, 실패하면 cp949(한국어 Windows
    엑셀의 기본 'CSV(쉼표로 분리)' 저장 인코딩) 로 재시도한다. 둘 다 실패하면 예외
    대신 사람이 읽을 오류로 돌려준다 - 골든셋 하나가 못 깨져도 CLI 가 죽으면 안 된다
    (이 모듈의 폴백 철학)."""
    for encoding in ("utf-8-sig", "cp949"):
        try:
            with open(path, encoding=encoding) as f:
                return f.read(), []
        except UnicodeDecodeError:
            continue
    return None, [f"{path}: 인코딩을 인식하지 못했습니다(utf-8/cp949 시도 실패)"]


def load_qa_set(path: str) -> tuple[dict[str, dict], list[str]]:
    """QA셋 파일 → {정규화 질문: {"ground_truth", "gold_contexts"}}, 오류 목록.

    JSON 배열 / JSONL / CSV / XLSX 를 받는다. 확장자로 1차 판별하되, JSON 파싱이
    실패하면 JSONL 로 재시도한다 - 상대가 .json 확장자로 JSONL 을 주는 실무 케이스."""
    ext = os.path.splitext(path)[1].lower()

    if ext in (".xlsx", ".xlsm"):
        entries, errors = _entries_from_xlsx(path)
    else:
        content, errors = _read_text(path)
        if content is None:
            return {}, errors
        if ext == ".csv":
            entries, errors = _entries_from_csv(content)
        elif ext == ".jsonl":
            entries, errors = _entries_from_jsonl(content)
        else:
            try:
                entries, errors = _entries_from_json(content)
            except json.JSONDecodeError:
                entries, errors = _entries_from_jsonl(content)

    qa_map: dict[str, dict] = {}
    for i, entry in enumerate(entries):
        if not isinstance(entry, dict):
            errors.append(f"항목 {i}: 오브젝트가 아님")
            continue
        question = _as_text(_first_key(entry, QUESTION_KEYS))
        if not question:
            errors.append(f"항목 {i}: 질문 없음")
            continue
        if normalize_question(question) in qa_map:
            errors.append(f"항목 {i}: 중복 질문(마지막 항목 우선): {question[:30]}")
        qa_map[normalize_question(question)] = {
            "ground_truth": _as_text(_first_key(entry, GROUND_TRUTH_KEYS)),
            "gold_contexts": _as_text_list(_first_key(entry, GOLD_CONTEXT_KEYS)),
        }
    return qa_map, errors


def merge_qa_into_log(log_path: str, qa_path: str, out_path: str) -> dict:
    """로그 JSONL 에 QA셋을 병합해 out_path 로 쓴다. 반환: 집계 dict."""
    qa_map, qa_errors = load_qa_set(qa_path)
    stats = {"log_lines": 0, "qa_entries": len(qa_map), "matched": 0,
             "filled_ground_truth": 0, "filled_gold_contexts": 0,
             "conflicts": 0, "qa_errors": qa_errors}
    matched_keys: set[str] = set()

    with open(log_path, encoding="utf-8-sig") as fin, \
            open(out_path, "w", encoding="utf-8") as fout:
        for line in fin:
            stripped = line.strip()
            if not stripped:
                continue
            stats["log_lines"] += 1
            try:
                obj = json.loads(stripped)
                assert isinstance(obj, dict)
            except (json.JSONDecodeError, AssertionError):
                fout.write(stripped + "\n")      # 깨진 줄은 그대로 통과(적재기가 집계)
                continue

            qa = qa_map.get(normalize_question(obj.get("question") or ""))
            if qa is not None:
                stats["matched"] += 1
                matched_keys.add(normalize_question(obj.get("question") or ""))
                if qa["ground_truth"]:
                    if not str(obj.get("ground_truth") or "").strip():
                        obj["ground_truth"] = qa["ground_truth"]
                        stats["filled_ground_truth"] += 1
                    elif str(obj["ground_truth"]).strip() != qa["ground_truth"]:
                        stats["conflicts"] += 1
                if qa["gold_contexts"]:
                    if not obj.get("gold_contexts"):
                        obj["gold_contexts"] = qa["gold_contexts"]
                        stats["filled_gold_contexts"] += 1
                    elif obj["gold_contexts"] != qa["gold_contexts"]:
                        stats["conflicts"] += 1     # GT 와 같은 규칙 - 불일치를 조용히 버리지 않는다
            fout.write(json.dumps(obj, ensure_ascii=False) + "\n")

    stats["qa_unmatched"] = len(qa_map) - len(matched_keys)
    return stats


# ── CLI ──────────────────────────────────────────────────────────

def _main(argv: list[str]) -> int:
    args = [a for a in argv if not a.startswith("--")]
    out = next((a.split("=", 1)[1] for a in argv if a.startswith("--out=")), None)
    if len(args) != 2:
        print("사용법: python -m agents.eval.qa_merge <log.jsonl> "
              "<qa.json|jsonl|csv|xlsx> [--out=경로]")
        return 2
    log_path, qa_path = args
    out = out or re.sub(r"\.jsonl?$", "", log_path) + ".merged.jsonl"

    stats = merge_qa_into_log(log_path, qa_path, out)
    print(f"로그 {stats['log_lines']}건 / QA셋 {stats['qa_entries']}건 "
          f"/ 매칭 {stats['matched']}건 (QA 미매칭 {stats['qa_unmatched']}건)")
    print(f"채움: 정답 {stats['filled_ground_truth']}건, "
          f"근거문단 {stats['filled_gold_contexts']}건 / 충돌(로그 값 유지) {stats['conflicts']}건")
    for err in stats["qa_errors"][:5]:
        print(f"  ! QA셋: {err}")
    print(f"출력: {out}")
    return 0 if stats["matched"] else 1


if __name__ == "__main__":
    # CLI 로 직접 부를 때만 콘솔 인코딩 보정(cp949 콘솔 한글 깨짐 방지).
    # 라이브러리 사용 시엔 core 의존이 없도록 여기서만 시도한다.
    try:
        from core.console import force_utf8_stdio
        force_utf8_stdio()
    except ImportError:
        pass
    sys.exit(_main(sys.argv[1:]))
